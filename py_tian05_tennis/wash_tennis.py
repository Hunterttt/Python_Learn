import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import openpyxl
import os,sys

os.chdir(sys.path[0]) 

sourse_html = 'http://www.benri.com/calendar/'      #根据网站不同链接不同

def get_syukujitsu():
    iheader = {"User-Agent": UserAgent().random}

    
    response = requests.get(sourse_html, headers = iheader)     #headers这个参数有些教程里没有

    #encodings = requests.utils.get_encodings_from_content(response.text)
    #print(encodings)
    #根据上面两行得到网站的编码后，把下面一行'UTF-8'替换成网站的编码，就不会出现乱码了
    ori_page = BeautifulSoup(response.content.decode('UTF-8'), 'lxml')      

    t_data1 = ori_page.find("div", {'id': 'ShukuList'})
    t_data2 = t_data1.find_all("div", {'class': 'SH_dt'})

    #print(t_data2)
    #print(type(t_data2))
    
    t_syukujitsu_list = []
    for i in range(1,len(t_data2)):       
        day_date = t_data2[i].text[5:-1]         #字符串中第5个位置开始的值
        
        t_syukujitsu_list.append(day_date.replace('月','/'))    #将字符串中的'月'替换为'/'
    
    return t_syukujitsu_list


def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)

def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]
    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组



def list2d_to_xlsx(xlist2d):
#写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save('washed_tennis.xlsx')


if __name__ == '__main__':
    syukujitsu_list = get_syukujitsu()
    print('今年の祝日,got from',sourse_html,'\n',syukujitsu_list)


    t_list2d = excel_to_list2d('Tennis.xlsx',0)

    #print(t_list2d[2][0])
    #print(type(t_list2d[2][0]))
    syukujitsu_tennis = []
    for i in range(len(t_list2d)):
        #print(type(t_list2d[i][0]))
        #print(str(t_list2d[i][0])[:-5])
        if '赤羽' in str(t_list2d[i][0]):        #可能是因为字符串里有'/',必须要用str()
            #print(str(t_list2d[i][0]))
            syukujitsu_tennis.append(t_list2d[i])

        elif '滝野川' in str(t_list2d[i][0]): 
            #print(str(t_list2d[i][0]))
            syukujitsu_tennis.append(t_list2d[i])

        elif t_list2d[i][0] == None:
            syukujitsu_tennis.append(t_list2d[i])

        elif '/' and '土' in t_list2d[i][0]:
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                else:
                    pass

        elif '/' and '日' in t_list2d[i][0]:
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                else:
                    pass

        elif str(t_list2d[i][0])[:-5] in syukujitsu_list:
            
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                else:
                    pass           

        else:
            pass
    #print(syukujitsu_tennis)
    list2d_to_xlsx(syukujitsu_tennis)

          



