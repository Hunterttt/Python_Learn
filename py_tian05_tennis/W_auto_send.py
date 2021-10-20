from contextlib import nullcontext
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import datetime
import openpyxl
import os,sys

import smtplib
from email.message import EmailMessage
import mimetypes



os.chdir(sys.path[0]) 

t = datetime.datetime.now()
itoday = t.strftime("%Y%m%d")
i7d = (datetime.datetime.now()+datetime.timedelta(days=7)).strftime("%Y%m%d")
i14d = (datetime.datetime.now()+datetime.timedelta(days=14)).strftime("%Y%m%d")
i21d = (datetime.datetime.now()+datetime.timedelta(days=21)).strftime("%Y%m%d")
i28d = (datetime.datetime.now()+datetime.timedelta(days=28)).strftime("%Y%m%d")


iheader = {"User-Agent": UserAgent().random}


def get_week1():
    '''
    写爬虫时会遇到提交表单的问题，一般先构造data，然后利用post方式进行提交表单。
    一般data的数据类型为字典，但当遇到多个数据项的属性名字重复时，则不能够使用字典了，因为字典中的键是不能够重复的。
    对于这个问题的解决可以使用，列表+元组的形式进行data数据的构造。如下：
    data=[('data_name','value'),('data_name','value'),('data_name','value')...]
    '''
    only1day = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_datetime_sel_1',
        data = [   
            ('g_sessionid', sessionid2),            
            ('chkbox', 'on'),
            ('u_yobi', 0),
            ('chkbox', 'on'),
            ('u_yobi', 1),
            ('chkbox', 'on'),
            ('u_yobi', 2),
            ('chkbox', 'on'),
            ('u_yobi', 3),
            ('chkbox', 'on'),
            ('u_yobi', 4),
            ('chkbox', 'on'),
            ('u_yobi', 5),
            ('chkbox', 'on'),
            ('u_yobi', 6),
            ('chkbox', 'on'),
            ('u_yobi', 10),
            ('ymd', itoday),
            ('u_genzai_idx', 6),
            ('g_kinonaiyo', 17)
        ],
        headers=iheader
    )
    
    week1 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_amenitytime_sel_1',
        data = {
            'g_sessionid': sessionid2,
            'flg_sstkoma': 0,
            'u_genzai_idx': 7,
            'g_kinonaiyo': 17,
            'showStartKoma': 1,
            'showEndKoma': 7
        }
    )
    return week1



def get_from_week2(yymmdd):
    from_week2 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_amenitytime_sel_1',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 7,
            'flg_ikou': 1,
            'hiduke_sousa_flg': 0,
            'u_hyojibi': yymmdd,
            'yoyakuinfo': None,        #空的时候用None
            #'g_kinonaiyo': 17,
            'showStartKoma': 1,
            'showEndKoma': 7
        }
    )
    return from_week2



def xotext_to_xolist(xotext):

    xotext_bs = BeautifulSoup(xotext.text, 'lxml')

    xotext_tr = xotext_bs.find_all('tr')

    for i in range(1,len(xotext_tr)):
        day_line = []    #最终表格中一个场地中一行的数据
       
        try:
        #在夏冬交替的时候，页面上的日期会出现空格
            day_date = xotext_tr[i].find('strong').text
        except:
            pass
            #day_date = "Placeholder"       #这个地方很奇怪，可能只有在一个页面上出现冬夏相交的时候才需要把pass换成这个值

        #print(day_date)
        day_line.append(day_date)    #最终表格中一个场地中一行的数据的开头是日期

        day_available = xotext_tr[i].find_all('td')

        #print(day_available)
        for day_interval in day_available:    #如果day_available为空，就没法执行这个循环，直接跳过

            if day_interval.find('img'):
            #当可以在网上直接预约的时候，显示的不是img，而是支持input的按钮
                img_line = day_interval.find('img')
            elif day_interval.find('input'):
                img_line = day_interval.find('input')
            else:
                img_line = ""     #这里必须给一个值，要不执行下面的语句的时候，img_line还是上一次循环的值
                #print(img_line)


            #在夏冬交替的时候，页面上的日期会出现空格
            try:
                alt = img_line.attrs.get('alt')
            except:
                #alt = "Placeholder"
                alt = ""
            day_line.append(alt)

        #print(day_line)
        list_1court.append(day_line)


def get_1court():
    
    week1 = get_week1()
    xotext_to_xolist(week1)
    try:
        week_temp = get_from_week2(i7d)
        xotext_to_xolist(week_temp)
    except:
        pass
    
    try:
        week_temp = get_from_week2(i14d)
        xotext_to_xolist(week_temp)
    except:
        pass
    
    try:
        week_temp = get_from_week2(i21d)
        xotext_to_xolist(week_temp)
    except:
        pass
    
    try:
        week_temp = get_from_week2(i28d)
        xotext_to_xolist(week_temp)
    except:
        pass

def list2d_to_xlsx(xlist2d):
#写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save('W_tennis.xlsx')


def add_list2d_to_xlsx(xlist2d,x,y):
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('W_tennis.xlsx')
    ws = wb['Sheet']
    # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
    for i in range(len(xlist2d)):
        for j in range(len(xlist2d[i])):
            ws.cell(row = x+i, column = y+j).value = xlist2d[i][j]
            # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
    # 保存操作
    wb.save('W_tennis.xlsx')






sourse_html = 'http://www.benri.com/calendar/'      #根据网站不同链接不同

def get_syukujitsu():
    iheader = {"User-Agent": UserAgent().random}

    
    response = requests.get(sourse_html, headers = iheader)     #headers这个参数有些教程里没有

    #encodings = requests.utils.get_encodings_from_content(response.text)
    #print(encodings)
    #根据上面两行得到网站的编码后，把下面一行'UTF-8'替换成网站的编码，就不会出现乱码了
    ori_page = BeautifulSoup(response.content.decode('UTF-8'), 'lxml')      

    t_data1 = ori_page.find("div", {'id': 'ShukuList'})
    t_data2 = t_data1.find_all("div", {'class': 'SH_dt'})

    #print(t_data2)
    #print(type(t_data2))
    
    t_syukujitsu_list = []
    for i in range(1,len(t_data2)):       
        day_date = t_data2[i].text[5:-1]         #字符串中第5个位置开始的值
        
        t_syukujitsu_list.append(day_date.replace('月','/'))    #将字符串中的'月'替换为'/'
    
    return t_syukujitsu_list


def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)

def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]
    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组



def list2d_to_xlsx_washed(xlist2d):
#写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save('W_washed_tennis.xlsx')



message = EmailMessage()

sender = 'tc@gmail.com'
receiver = 'ttt@hotmail.com'
pswd = 'wa5'

def send_email(xflie):
    message['From'] = sender
    message['To'] = receiver
    message['Subject'] = 'Subject line goes here'

    body_of_email = 'Text to be displayed in the email'
    message.set_content(body_of_email)


    mime_type, _ = mimetypes.guess_type(xflie)           #不懂
    mime_type, mime_subtype = mime_type.split('/')           #不懂
    with open(xflie, 'rb') as file:
        message.add_attachment(
        file.read(),
        maintype = mime_type,
        subtype = mime_subtype,
        filename = xflie)
    #print(message)


    mail_server = smtplib.SMTP_SSL(host = 'smtp.gmail.com', port = 465)
    #mail_server.set_debuglevel(1)
    mail_server.login(user = sender, password = pswd)
    mail_server.send_message(message)
    mail_server.quit()




if __name__ == "__main__":

    session1 = requests.session()

    page1 = session1.get('https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_menu',headers=iheader)
    page1_bs = BeautifulSoup(page1.text, 'lxml')

    #print(page1.text)
    sessiont1 = page1_bs.find(attrs={'name':'RiyosyaForm'})
    sessiont2 = sessiont1.find(attrs={'name':'g_sessionid'})
    sessionid1 = sessiont2.attrs.get('value')

    #print(sessionid1)


    page2 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_init',
        data = {
            'g_sessionid': sessionid1,
            'x': 77,
            'y': 44,
            'u_genzai_idx': 0,
            'g_kinonaiyo': 35
        },
        headers=iheader
    )
    #print(page2.text)


    page3 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_id_input',
        data = {
            'g_sessionid': sessionid1,
            'x': 337,
            'y': 63,
            'u_genzai_idx': 0,
            'g_kinonaiyo': 35
        },
        headers=iheader
    )
    #print(page3.text)



    page4 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_login1',
        data = {
            'g_sessionid': sessionid1,
            'u_genzai_idx': 1,
            'g_kinonaiyo': 14,
            'g_riyoushabangou': 3100019255,
            'ansyono': 3385,
            'x': 79,
            'y': 42
        },
        headers=iheader
    )

    page4_bs = BeautifulSoup(page4.text, 'lxml')
    sessiont3 = page4_bs.find(attrs={'name':'YykForm'})
    sessiont4 = sessiont3.find(attrs={'name':'g_sessionid'})
    sessionid2 = sessiont4.attrs.get('value')

    #print(sessionid2)


    page5 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_group_sel_1',
        data = {
            'g_sessionid': sessionid2,
            'x': 107,
            'y': 73,
            'u_genzai_idx': 0,
            'g_kinonaiyo': 17
        },
        headers=iheader
    )
    #print(page5.text)



    page6 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_group_sel_2',
        data = {
            'g_sessionid': sessionid2,
            'checkedAll': 'false',
            'g_bunruicd': 5001,
            'bunruicd': 5001,
            'u_genzai_idx': 1,
            'g_kinonaiyo': 17,
            'pageflg': 1
        },
        headers=iheader
    )
    #print(page6.text)


    page7 = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dest_sel',
        data = {
            'g_sessionid': sessionid2,
            'checkedAll': 'false',
            'g_bunruicd': 5101,
            'bunruicd': 5101,
            'u_genzai_idx': 2,
            'g_kinonaiyo': 17,
            'pageflg': 2
        },
        headers=iheader
    )
    #print(page7.text)





    '''
    显示所有地方的页面
    '''
    page_courts = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_amenity_sel',
        data = {
            'g_sessionid': sessionid2,
            'riyosmk': 5000,
            'u_genzai_idx': 3,
            'g_kinonaiyo': 17,
        },
        headers=iheader
    )
    #print(page_courts.text)

    '''
    显示所有球场的页面
    '''
    akabane_kiri = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_room_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 4,
            'g_kinonaiyo': 17,
            'g_basyocd': 63,
            'g_jitubasyocd': 521,
            'shisetugroup': 50,
            'g_systemcd': 1,
            'g_mkkbn': 2
        },
        headers=iheader
    )
    #print(akabane_kiri.text)


    

    '''
    赤羽，桐丘A面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5056100
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘A面')
    list_1court = []
    list_1court.append(['赤羽，桐丘A面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    list2d_to_xlsx(list_1court)

    '''
    赤羽，桐丘B面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5056200
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘B面')
    list_1court = []
    list_1court.append(['赤羽，桐丘B面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,1,9)


    '''
    赤羽，桐丘C面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5056300
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘C面')
    list_1court = []
    list_1court.append(['赤羽，桐丘C面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,1,17)


    '''
    赤羽，桐丘D面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5056400
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘D面')
    list_1court = []
    list_1court.append(['赤羽，桐丘D面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,1,25)


    '''
    赤羽，桐丘E面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5055100
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘E面')
    list_1court = []
    list_1court.append(['赤羽，桐丘E面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,1,33)


    '''
    赤羽，桐丘F面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 63,
            'g_stgroup': 50,
            'g_sisetucd': 5055200
        },
        headers=iheader
    )
    print('checking 赤羽，桐丘F面')
    list_1court = []
    list_1court.append(['赤羽，桐丘F面'])
    list_1court.append(['','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,1,41)









    '''
    显示所有地方的页面
    '''
    page_courts = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_amenity_sel',
        data = {
            'g_sessionid': sessionid2,
            'riyosmk': 5000,
            'u_genzai_idx': 3,
            'g_kinonaiyo': 17,
        },
        headers=iheader
    )
    #print(page_courts.text)

    '''
    显示所有球场的页面
    '''
    akabane_kiri = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_room_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 4,
            'g_kinonaiyo': 17,
            'g_basyocd': 66,
            'g_jitubasyocd': 532,
            'shisetugroup': 50,
            'g_systemcd': 1,
            'g_mkkbn': 2
        },
        headers=iheader
    )
    #print(akabane_kiri.text)

    '''
    赤羽，河岸A面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 66,
            'g_stgroup': 50,
            'g_sisetucd': 5070100
        },
        headers=iheader
    )
    print('checking 赤羽，河岸A面')
    list_1court = []
    list_1court.append(['赤羽，河岸A面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00'])
    get_1court()
    #print(list_1court)
    add_list2d_to_xlsx(list_1court,41,1)


    '''
    赤羽，河岸B面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 66,
            'g_stgroup': 50,
            'g_sisetucd': 5070200
        },
        headers=iheader
    )
    print('checking 赤羽，河岸B面')
    list_1court = []
    list_1court.append(['赤羽，河岸B面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,41,9)


    '''
    赤羽，河岸C面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 66,
            'g_stgroup': 50,
            'g_sisetucd': 5070300
        },
        headers=iheader
    )
    print('checking 赤羽，河岸C面')
    list_1court = []
    list_1court.append(['赤羽，河岸C面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,41,17)


    '''
    赤羽，河岸D面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 66,
            'g_stgroup': 50,
            'g_sisetucd': 5070400
        },
        headers=iheader
    )
    print('checking 赤羽，河岸D面')
    list_1court = []
    list_1court.append(['赤羽，河岸D面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,41,25)


    '''
    赤羽，河岸E面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 66,
            'g_stgroup': 50,
            'g_sisetucd': 5070500
        },
        headers=iheader
    )
    print('checking 赤羽，河岸E面')
    list_1court = []
    list_1court.append(['赤羽，河岸E面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,41,33)







    '''
    显示所有地方的页面
    '''
    page_courts = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_amenity_sel',
        data = {
            'g_sessionid': sessionid2,
            'riyosmk': 5000,
            'u_genzai_idx': 3,
            'g_kinonaiyo': 17,
        },
        headers=iheader
    )
    #print(page_courts.text)

    '''
    显示所有球场的页面
    '''
    akabane_kiri = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_room_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 4,
            'g_kinonaiyo': 17,
            'g_basyocd': 61,
            'g_jitubasyocd': 511,
            'shisetugroup': 50,
            'g_systemcd': 1,
            'g_mkkbn': 2
        },
        headers=iheader
    )
    #print(akabane_kiri.text)

    '''
    滝野川A面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 61,
            'g_stgroup': 50,
            'g_sisetucd': 5045100
        },
        headers=iheader
    )
    print('checking 滝野川A面')
    list_1court = []
    list_1court.append(['滝野川A面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,81,1)


    '''
    滝野川B面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 61,
            'g_stgroup': 50,
            'g_sisetucd': 5045200
        },
        headers=iheader
    )
    print('checking 滝野川B面')
    list_1court = []
    list_1court.append(['滝野川B面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,81,10)


    '''
    滝野川C面
    '''
    akabane_kiri_date_sel = session1.post(
        url = 'https://yoyaku.city.kita.tokyo.jp/shisetsu/reserve/gin_z_dsp_sel',
        data = {
            'g_sessionid': sessionid2,
            'u_genzai_idx': 5,
            'g_kinonaiyo': 17,
            'g_kinomkkbn': 2,
            'g_basyocd': 61,
            'g_stgroup': 50,
            'g_sisetucd': 5045300
        },
        headers=iheader
    )
    print('checking 滝野川C面')
    list_1court = []
    list_1court.append(['滝野川C面'])
    list_1court.append(['','07:00~09:00','09:00~11:00','11:00~13:00','13:00~15:00','15:00~17:00','18:00~20:00'])
    get_1court()
    add_list2d_to_xlsx(list_1court,81,19)







    syukujitsu_list = get_syukujitsu()
    print('今年の祝日,got from',sourse_html,'\n',syukujitsu_list)


    t_list2d = excel_to_list2d('W_tennis.xlsx',0)

    #print(t_list2d[2][0])
    #print(type(t_list2d[2][0]))
    syukujitsu_tennis = []
    num_of_cournt = 0

    for i in range(len(t_list2d)):
        #print(type(t_list2d[i][0]))
        #print(str(t_list2d[i][0])[:-5])
        if '赤羽' in str(t_list2d[i][0]):        #可能是因为字符串里有'/',必须要用str()
            #print(str(t_list2d[i][0]))
            syukujitsu_tennis.append(t_list2d[i])

        elif '滝野川' in str(t_list2d[i][0]): 
            #print(str(t_list2d[i][0]))
            syukujitsu_tennis.append(t_list2d[i])

        elif t_list2d[i][0] == None:
            syukujitsu_tennis.append(t_list2d[i])

        elif '/' and '土' in t_list2d[i][0]:
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                else:
                    pass

        elif '/' and '日' in t_list2d[i][0]:
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                else:
                    pass

        elif str(t_list2d[i][0])[:-5] in syukujitsu_list:    #祝日的情况
            
            for j in range(len(t_list2d[i])):
                if 'O' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                elif '問' in str(t_list2d[i][j]):
                    syukujitsu_tennis.append(t_list2d[i])
                    num_of_cournt += 1
                    break
                else:
                    pass           

        else:
            pass
    #print(syukujitsu_tennis)
    list2d_to_xlsx_washed(syukujitsu_tennis)
    if num_of_cournt > 0:
        send_email('W_washed_tennis.xlsx')
    else:
        pass



